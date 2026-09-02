#!/usr/bin/env python3
"""Holt Tennisdaten.

Ergebnisse und Quoten von tennis-data.co.uk (ATP und WTA, je Satz die
Spielstaende, dazu Belag, Runde, Best-of, Weltranglistenplaetze und Quoten).

Sackmanns Repos tennis_atp und tennis_wta gibt es nicht mehr; damit entfaellt
die einzige freie Quelle fuer Aufschlagstatistiken je Match. Das Modell setzt
deshalb auf gewonnenen Spielen auf, nicht auf Punkten.

Ansetzungen ueber die offene Scoreboard-Schnittstelle von ESPN.

Erzeugt:
  daten/tennis-atp.csv
  daten/tennis-wta.csv
  daten/tennis-ansetzungen.csv
"""

import csv, io, json, os, time, urllib.error, urllib.request

JAHRE = [2024, 2025, 2026]
ORDNER = "daten"
BASIS = "http://www.tennis-data.co.uk"

# Die Seite bietet je nach Jahrgang csv oder xlsx an; beides wird probiert.
MUSTER = {
    "atp": ["{j}/{j}.csv", "{j}/{j}.xlsx"],
    "wta": ["{j}w/{j}.csv", "{j}w/{j}.xlsx"],
}

KOPF = ["Datum","Turnier","Serie","Court","Belag","Runde","BestOf",
        "Sieger","Verlierer","SRang","VRang",
        "S1","V1","S2","V2","S3","V3","S4","V4","S5","V5",
        "SSaetze","VSaetze","SSpiele","VSpiele","Status","QuoteS","QuoteV"]


def hole(url, versuche=2):
    letzter = "unbekannt"
    for i in range(versuche):
        try:
            req = urllib.request.Request(url, headers={
                "User-Agent": "Mozilla/5.0 (compatible; linien-modell/1.0)",
                "Accept": "*/*",
            })
            with urllib.request.urlopen(req, timeout=60) as r:
                return r.read(), "ok"
        except urllib.error.HTTPError as e:
            return None, f"HTTP {e.code}"
        except Exception as e:
            letzter = type(e).__name__
            if i < versuche - 1:
                time.sleep(2)
    return None, letzter


def zahl(v):
    try:
        x = float(str(v).strip())
        return x
    except Exception:
        return None


def ganz(v):
    x = zahl(v)
    return int(x) if x is not None else None


def zeilen_aus_csv(rohdaten):
    for kodierung in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            text = rohdaten.decode(kodierung)
            break
        except Exception:
            continue
    else:
        return []
    return list(csv.DictReader(io.StringIO(text)))


def zeilen_aus_xlsx(rohdaten):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("    openpyxl fehlt, xlsx wird uebersprungen")
        return []
    try:
        wb = load_workbook(io.BytesIO(rohdaten), read_only=True, data_only=True)
        blatt = wb[wb.sheetnames[0]]
        reihen = blatt.iter_rows(values_only=True)
        kopf = [str(k).strip() if k is not None else "" for k in next(reihen)]
        raus = []
        for r in reihen:
            if r is None or all(x is None for x in r):
                continue
            raus.append({kopf[i]: r[i] for i in range(min(len(kopf), len(r)))})
        return raus
    except Exception as e:
        print(f"    xlsx nicht lesbar: {type(e).__name__}")
        return []


def datum_text(v):
    if v is None:
        return ""
    s = str(v)
    return s[:10] if "-" in s or "/" in s else s


def erste_quote(z, felder):
    for feld in felder:
        x = zahl(z.get(feld))
        if x and x > 1:
            return x
    return ""


def verarbeiten(zeilen):
    raus = []
    for z in zeilen:
        sieger = str(z.get("Winner") or "").strip()
        verlierer = str(z.get("Loser") or "").strip()
        if not sieger or not verlierer:
            continue
        saetze = []
        s_spiele = v_spiele = 0
        gueltig = False
        for i in range(1, 6):
            a, b = ganz(z.get(f"W{i}")), ganz(z.get(f"L{i}"))
            saetze += [a if a is not None else "", b if b is not None else ""]
            if a is not None and b is not None:
                s_spiele += a
                v_spiele += b
                gueltig = True
        if not gueltig:
            continue
        raus.append([
            datum_text(z.get("Date")),
            str(z.get("Tournament") or "").strip(),
            str(z.get("Series") or z.get("Tier") or "").strip(),
            str(z.get("Court") or "").strip(),
            str(z.get("Surface") or "").strip(),
            str(z.get("Round") or "").strip(),
            ganz(z.get("Best of")) or "",
            sieger, verlierer,
            ganz(z.get("WRank")) or "", ganz(z.get("LRank")) or "",
        ] + saetze + [
            ganz(z.get("Wsets")) or "", ganz(z.get("Lsets")) or "",
            s_spiele, v_spiele,
            str(z.get("Comment") or "").strip(),
            erste_quote(z, ["AvgW", "PSW", "B365W", "MaxW"]),
            erste_quote(z, ["AvgL", "PSL", "B365L", "MaxL"]),
        ])
    return raus


def tour_holen(tour):
    alle = []
    for jahr in JAHRE:
        geschafft = False
        for m in MUSTER[tour]:
            name = m.format(j=jahr)
            roh, hinweis = hole(f"{BASIS}/{name}")
            if not roh or len(roh) < 500:
                print(f"  fehlt {name}  ({hinweis})")
                continue
            zeilen = (zeilen_aus_xlsx(roh) if name.endswith(".xlsx")
                      else zeilen_aus_csv(roh))
            fertig = verarbeiten(zeilen)
            if not fertig:
                print(f"  leer  {name}")
                continue
            alle += fertig
            print(f"  ok    {name}  {len(fertig)} Partien")
            geschafft = True
            break
        if not geschafft:
            print(f"  Jahr {jahr} fuer {tour.upper()} nicht geholt")
        time.sleep(0.3)
    return alle


def ansetzungen_holen():
    raus = []
    gesehen = set()
    for tour in ("atp", "wta"):
        url = (f"https://site.api.espn.com/apis/site/v2/sports/tennis/"
               f"{tour}/scoreboard")
        roh, hinweis = hole(url)
        if not roh:
            print(f"  ESPN {tour}: {hinweis}")
            continue
        try:
            d = json.loads(roh.decode("utf-8", errors="replace"))
        except Exception:
            print(f"  ESPN {tour}: Antwort war kein JSON")
            continue
        anz = 0
        for ev in d.get("events", []):
            turnier = ev.get("name", "")
            for gr in ev.get("groupings", []):
                for wett in gr.get("competitions", []):
                    zust = (wett.get("status", {})
                                .get("type", {}).get("state", ""))
                    if zust == "post":
                        continue
                    leute = wett.get("competitors", [])
                    if len(leute) != 2:
                        continue
                    namen = []
                    for k in leute:
                        a = k.get("athlete") or {}
                        namen.append((a.get("displayName")
                                      or a.get("shortName") or "").strip())
                    if not all(namen):
                        continue
                    schluessel = (turnier, namen[0], namen[1])
                    if schluessel in gesehen:
                        continue
                    gesehen.add(schluessel)
                    raus.append([
                        tour.upper(), wett.get("date", ""), turnier,
                        wett.get("format", {}).get("bestOf", ""),
                        gr.get("grouping", {}).get("shortName", ""),
                        namen[0], namen[1],
                    ])
                    anz += 1
        print(f"  ESPN {tour}: {anz} neue Partien")
    return raus


def schreiben(pfad, kopf, zeilen):
    with open(pfad, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(kopf)
        w.writerows(zeilen)
    print(f"{pfad}: {len(zeilen)} Zeilen")


def main():
    os.makedirs(ORDNER, exist_ok=True)
    for tour in ("atp", "wta"):
        print(f"{tour.upper()} holen ...")
        schreiben(f"{ORDNER}/tennis-{tour}.csv", KOPF, tour_holen(tour))

    print("Ansetzungen holen ...")
    schreiben(f"{ORDNER}/tennis-ansetzungen.csv",
              ["Tour", "Zeit", "Turnier", "BestOf", "Runde", "Spieler1", "Spieler2"],
              ansetzungen_holen())


if __name__ == "__main__":
    main()
